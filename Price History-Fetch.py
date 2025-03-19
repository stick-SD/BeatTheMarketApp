import os
import pandas as pd
from yahoo_fin import stock_info as si
from openpyxl import load_workbook
import shutil
from openpyxl.styles import Alignment
from datetime import datetime
import time
import sys
 
def fetch_and_update_ticker(ticker):
    try:
        #print(f"Fetching data for {ticker}...")
        stock_data = si.get_data(ticker)

        # Check if stock_data is empty
        if stock_data.empty:
            print(f"No data available for {ticker}.")
            return

        # Format stock_data's Date column (which is currently the index)
        stock_data.index = stock_data.index.strftime('%m/%d/%Y')
        stock_data.reset_index(inplace=True)
        stock_data.rename(columns={'index': 'Date'}, inplace=True)

        # Drop the 'ticker' column if it exists
        if 'ticker' in stock_data.columns:
            stock_data.drop(columns=['ticker'], inplace=True)

        # Special handling for ^GSPC and ^DJI (handling the pre-existing data)
        pre_existing_file = None
        if ticker == "^GSPC":
            pre_existing_file = "^GSPC_Pre_Mar-25-1970.xlsx"
            sheet_name = "PriceData"
        elif ticker == "^DJI":
            pre_existing_file = "^DJI_Pre_Jan-2-1992.xlsx"
            sheet_name = "PriceData"

        if pre_existing_file and os.path.exists(pre_existing_file):
            # Load pre-existing data and ensure it has a 'Date' column
            pre_existing_data = pd.read_excel(pre_existing_file, sheet_name=sheet_name, index_col=0)
            pre_existing_data.index = pd.to_datetime(pre_existing_data.index)

            # Reset the index to move the date into a 'Date' column like in stock_data
            pre_existing_data.reset_index(inplace=True)
            pre_existing_data.rename(columns={'index': 'Date'}, inplace=True)

            # Ensure the 'Date' column is formatted the same way as in stock_data
            pre_existing_data['Date'] = pre_existing_data['Date'].dt.strftime('%m/%d/%Y')

            # Rename pre_existing_data columns to match stock_data
            pre_existing_data.rename(columns={'Open': 'open', 'High': 'high', 'Low': 'low', 
                                              'Close': 'close', 'AdjClose': 'adjclose', 'Volume': 'volume'}, inplace=True)

            # Combine the pre-existing data with stock_data
            combined_data = pd.concat([pre_existing_data, stock_data])

            # Remove duplicate dates, keeping the first occurrence
            combined_data = combined_data[~combined_data['Date'].duplicated(keep='first')]
        else:
            combined_data = stock_data

        # Define the output directory for ticker data files
        output_directory = "Ticker Data"
        
        # Create the directory if it doesn't exist
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)
            print(f"Created directory: {output_directory}")
        
        # Define the output file path within the "Ticker Data" directory
        output_file = os.path.join(output_directory, f"{ticker}.xlsx")

        template_file = "Price History Template.xlsx"

        # Check if the file exists; if not, copy the template file
        if not os.path.exists(output_file):
            shutil.copy(template_file, output_file)
            #print(f"Created new file from template: {output_file}")

        # Load the workbook and select the sheet
        wb = load_workbook(output_file)
        if 'PriceData' not in wb.sheetnames:
            raise ValueError("Sheet named 'PriceData' does not exist in the workbook.")

        sheet = wb['PriceData']

        # Prepare list to store data rows
        data = []

        # Loop over rows of the combined data and append to the list
        for index, row in combined_data.iterrows():
            date = row['Date']

            # Check if the date is valid, otherwise log error and skip this row
            if pd.isna(date):
                print(f"Error: Invalid date encountered at row {index}. Skipping this row.")
                continue

            open_price = row['open']
            high_price = row['high']
            low_price = row['low']
            close_price = row['close']
            adj_close = row['adjclose']
            volume = row['volume']

            # Append data row to the list
            data.append([date, open_price, high_price, low_price, close_price, adj_close, volume])

        # Define the headers
        headers = ["Date", "Open", "High", "Low", "Close", "AdjClose", "Volume"]

        # Clear previous content, but preserve formatting
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=len(headers), values_only=False):
           for cell in row:
                cell.value = None  # Clear the content, but keep the formatting

        # Write headers
        for col_num, column_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=column_title)
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Center the headers

        # Write new data starting from the second row
        for i, row_data in enumerate(data, start=2):
            for j, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=i, column=j)
                if j == 1:  # Assuming first column is 'Date'
                    try:
                        date_value = datetime.strptime(value, '%m/%d/%Y')
                        cell.value = date_value
                        cell.number_format = 'mm/dd/yyyy'
                    except Exception as e:
                        print(f"Error parsing date '{value}': {e}")
                        continue
                elif j >= 2 and j <= 6:  # Format numerical columns (Open, High, Low, Close, AdjClose)
                    cell.value = value
                    cell.number_format = '0.00'  # Format with two decimal places
                else:
                    cell.value = value

                # Center the data cells
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Save the workbook
        wb.save(output_file)

        # Get today's date and the last date of data
        today = datetime.now().strftime('%m/%d/%Y')
        last_date = stock_data.iloc[-1]['Date']
        
        # Print updated message with today's date and the last date of data
        print(f"Ticker {ticker} data updated. Last date of data {last_date}.")

    except Exception as e:
        print(f"Error fetching data for {ticker}: {e}")


def read_ticker_list(file_name):
    try:
        with open(file_name, 'r') as f:
            tickers = [line.strip() for line in f if line.strip()]
        return tickers
    except Exception as e:
        print(f"Error reading ticker list: {e}")
        return []


def main():
    if len(sys.argv) < 2:
        print("Error: No ticker list file provided.")
        return

    # Get the ticker file from command-line arguments
    ticker_file = sys.argv[1]

    # Check if the file exists
    if not os.path.exists(ticker_file):
        print(f"Error: File {ticker_file} does not exist.")
        return

    # Read tickers from text file in the current directory
    tickers = read_ticker_list(ticker_file)
    print(f"Starting ticker download from file {ticker_file}.")

    if not tickers:
        print(f"No tickers found in {ticker_file}. Exiting...")
        return

    # Fetch and update data for each ticker
    for ticker in tickers:
        fetch_and_update_ticker(ticker)
        time.sleep(2)  # Add a delay of 2 seconds between requests

    #input("Press enter to close this window...")
    print("Done.")


if __name__ == "__main__":
    main()