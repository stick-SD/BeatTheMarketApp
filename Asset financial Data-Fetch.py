import yfinance as yf
import pandas as pd
import os
import re
from openpyxl import load_workbook

# Function to read tickers from a CSV file
def load_tickers(file_path):
    if os.path.exists(file_path):
        print(f"Reading tickers from {file_path}...")
        try:
            # Read the ticker list from the CSV file
            return pd.read_csv(file_path, header=None)[0].tolist()
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            return []  # Return an empty list in case of error
    else:
        print(f"{file_path} not found.")
        return []  # Return an empty list if the file is not found

def fetch_data(tickers):
    if not tickers:
        print("No tickers to fetch data for. Exiting.")
        return

    print("Starting data fetch...")
    data = []

    # Initialize the error log file
    error_log_file = "Asset Financial Data-ErrorLog.txt"
    if os.path.exists(error_log_file):
        os.remove(error_log_file)  # Remove the existing log file to start fresh

    # Loop through each ticker and gather information
    for ticker in tickers:
        try:
            #print(f"Fetching data for {ticker}...")
            stock = yf.Ticker(ticker)
            info = stock.info

            company_name = info.get("longName", "N/A")
            market_cap = info.get("marketCap", "N/A")

            # Ensure market_cap is a number before dividing
            if isinstance(market_cap, (int, float)):
                market_cap = market_cap / 1000000  # Convert to millions
            else:
                market_cap = "N/A"

            trailing_pe_TTM = info.get("trailingPE", "N/A")
            forward_pe = info.get('forwardPE', "N/A")
            earnings_growth_rate = info.get('earningsGrowth', "N/A")
            peg_ratio = info.get('pegRatio', "N/A")

            trailingAnnualDividendYield= info.get("trailingAnnualDividendYield", "N/A")
            forward_dividend_yield = info.get("dividendYield", "N/A")  # This is forward dividend yield for Equities
            forward_yield = info.get("yield", "N/A")  # This is forward dividend yield for Funds

            # If forward_dividend_yield is "N/A", set it equal to forward_yield
            if forward_dividend_yield == "N/A":
                forward_dividend_yield = forward_yield

            payout_ratio = info.get("payoutRatio", "N/A")

            price_sales_ratio_TTM = info.get('priceToSalesTrailing12Months', "N/A")
            price_book_ratio = info.get('priceToBook', "N/A")

            beta_3_year = info.get("beta3Year", "N/A")
            beta_5y_monthly = info.get('beta', "N/A")

            fund_expense_ratio = info.get("annualReportExpenseRatio", "N/A")

            asset_type = info.get('quoteType', "N/A")
            category = info.get('category', "N/A")
            sector = info.get("sector", "N/A")
            industry = info.get("industry", "N/A")
            company_profile = info.get("longBusinessSummary", "N/A")

            book_value = info.get("bookValue", "N/A")
            trailing_peg_ratio = info.get("trailingPegRatio", "N/A")
            total_debt = info.get("totalDebt", "N/A")
            ytd_return = info.get("ytdReturn", "N/A")
            fifty_two_week_change = info.get("52WeekChange", "N/A")
            three_year_avg_return = info.get("threeYearAverageReturn", "N/A")
            five_year_avg_return = info.get("fiveYearAverageReturn", "N/A")

            # Initialize variables
            assetType = "N/A"
            subAssetType1 = "N/A"
            subAssetType2 = "N/A"
            subAssetType3 = "N/A"
            subAssetType4 = "N/A"
            subAssetType5 = "N/A"

            # Search for "closed end" or "closed-end" (case-insensitive)
            if re.search(r"closed[-\s]?end", company_profile, re.IGNORECASE):
                subAssetType1 = "CEF"

            # Search for "Corporation" (case-insensitive)
            if re.search(r"corporation", company_profile, re.IGNORECASE):
                subAssetType2 = "Corp"

            # Search for "Limited Liability" or "LLC" (case-insensitive)
            if re.search(r"limited liability|LLC", company_profile, re.IGNORECASE):
                subAssetType3 = "LLC"

            # Search for "business development company" (case-insensitive)
            if re.search(r"business development company|BDC", company_profile, re.IGNORECASE):
                subAssetType4 = "BDC"

            # Search for "master limited" (case-insensitive)
            if re.search(r"master limited|MLP", company_profile, re.IGNORECASE):
                subAssetType5 = "MLP"

            # Combine all subAssetType parameters that are not "N/A"
            assetType_list = [asset_type, subAssetType1, subAssetType2, subAssetType3, subAssetType4, subAssetType5, category, sector, industry]
            assetType = ', '.join([t for t in assetType_list if t != "N/A"])

            # Print the results
            #print(f"Asset Type: {assetType}")

            # Append the data to the list
            data.append([
                company_name, ticker, market_cap, trailing_pe_TTM, forward_pe, earnings_growth_rate, peg_ratio, 
                trailingAnnualDividendYield, forward_dividend_yield, payout_ratio, price_sales_ratio_TTM, price_book_ratio, 
                beta_3_year, beta_5y_monthly, fund_expense_ratio, asset_type, category, sector, industry, assetType, company_profile
            ])

            print(f"Data for {ticker} fetched successfully.")
            
        except Exception as e:
            error_message = f"Error fetching data for {ticker}: {str(e)}\n"
            print(error_message)
            
            # Write the error to the log file
            with open(error_log_file, "a") as log_file:
                log_file.write(error_message)

    if not data:
        print("No data fetched. Exiting.")
        return

    # Define the columns
    columns = [
        "Company Name", "Ticker", "Market Cap, Millions", "Trailing PE Ratio (TTM)", "PE Ratio (Fwd)", "Earning Growth Rate", "PEG Ratio",
        "Dividend Yield (TTM)", "Dividend Yield (Fwd)", "Payout Ratio", "Price to Sales Ratio (TTM)", "Price to Book Ratio", 
        "Beta 3yr", "Beta 5yr Monthly", "Fund Expense Ratio", "Asset Type", "Category", "Sector", "Industry", "Asset Type Summary", "Company Profile"
    ]

    # Load the existing workbook without overwriting it
    output_file = "Asset Financial Data-Results.xlsx"
    try:
        if os.path.exists(output_file):
            #print(f"Opening existing file: {output_file}")
            wb = load_workbook(output_file)
            ws = wb['FinancialData'] if 'FinancialData' in wb.sheetnames else wb.create_sheet('FinancialData')

            # Clear previous content, but preserve formatting
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(columns), values_only=False):
                for cell in row:
                    cell.value = None  # Clear the content, but keep the formatting

            # Write headers
            for col_num, column_title in enumerate(columns, 1):
                ws.cell(row=1, column=col_num, value=column_title)

            # Write new data starting from the second row
            for i, row_data in enumerate(data, start=2):
                for j, value in enumerate(row_data, start=1):
                    ws.cell(row=i, column=j, value=value)

        else:
            print(f"Creating new file: {output_file}")
            # Create a new workbook if file doesn't exist
            wb = load_workbook(output_file)
            ws = wb.create_sheet('FinancialData')

            # Write headers
            for col_num, column_title in enumerate(columns, 1):
                ws.cell(row=1, column=col_num, value=column_title)

            # Write new data
            for i, row_data in enumerate(data, start=2):
                for j, value in enumerate(row_data, start=1):
                    ws.cell(row=i, column=j, value=value)

        # Save the updated file
        wb.save(output_file)
        print(f"Data saved to {output_file}")

    except Exception as e:
        error_message = f"Error saving data to Excel: {str(e)}\n"
        print(error_message)
        with open(error_log_file, "a") as log_file:
            log_file.write(error_message)

# Prompt the user to select which ticker list to load
def select_ticker_list():
    while True:
        print("\nPlease choose a ticker list to load:")
        print("1: TickerList_1.csv")
        print("2: TickerList_2.csv")
        choice = input("Enter the number of your choice: ")

        if choice == '1':
            return "TickerList_1.csv"
        elif choice == '2':
            return "TickerList_2.csv"
        else:
            print("Invalid choice. Please try again.")

# Main logic
ticker_file = select_ticker_list()
tickers = load_tickers(ticker_file)

# Fetch and process data for the selected tickers
fetch_data(tickers)